"""
Step 1: Parse the xlsx scam datasets into a unified JSONL format.

Output format per line:
{
    "source": "sms" | "url" | "social_media",
    "input_text": "...",
    "label": "scam"
}
"""

import json
import os
import openpyxl
from config import (
    STAGE2_RAW_DIR as DATASET_DIR,
    SMS_URL_FILE, SOCIAL_FILE as SOCIAL_MEDIA_FILE,
    STAGE2_OUT_DIR as OUTPUT_DIR, S2_RAW_SCAM as RAW_SCAM_JSON,
)


def parse_sms_url(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath)
    records = []

    # SMS sheet
    ws = wb["Scam SMS dataset"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # skip header
        text = row[0]
        if text and str(text).strip():
            records.append({
                "source": "sms",
                "input_text": str(text).strip(),
                "label": "scam",
            })

    # URL sheet
    ws = wb["Scam URL dataset"]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        url = row[0]
        if url and str(url).strip():
            records.append({
                "source": "url",
                "input_text": str(url).strip(),
                "label": "scam",
            })

    return records


def parse_social_media(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Data"]
    records = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        content = row[1]  # column B = content
        if content and str(content).strip():
            records.append({
                "source": "social_media",
                "input_text": str(content).strip(),
                "label": "scam",
            })

    return records


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    sms_url_path = os.path.join(DATASET_DIR, SMS_URL_FILE)
    social_path = os.path.join(DATASET_DIR, SOCIAL_MEDIA_FILE)

    records = []
    records.extend(parse_sms_url(sms_url_path))
    records.extend(parse_social_media(social_path))

    with open(RAW_SCAM_JSON, "w", encoding="utf-8") as f:
        for rec in records:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    # Stats
    sources = {}
    for r in records:
        sources[r["source"]] = sources.get(r["source"], 0) + 1

    print(f"✅ Parsed {len(records)} scam records:")
    for src, cnt in sources.items():
        print(f"   {src}: {cnt}")
    print(f"   Output: {RAW_SCAM_JSON}")


if __name__ == "__main__":
    main()
